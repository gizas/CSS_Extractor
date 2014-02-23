from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import NoAlertPresentException
import unittest, time, re
import logging

#Include Excel functions for saving data in excel
from openpyxl import Workbook,load_workbook
from openpyxl.style import Color, Fill
col1='A'  #For Excel Manipulation of Cells -------------
col2='B' #For Excel Manipulation of Cells -------------


class test1(unittest.TestCase):
    def setUp(self):
        FORMAT = "%(message)-s"
        logging.basicConfig(format=FORMAT,filename='reportExternalCSS.txt', filemode='a', level=logging.INFO)
        self.logger2 = logging.getLogger("f2")#Create second logger
        self.handler2 = logging.FileHandler('reportExternalCSS.txt',mode='a')#Create Filehandler
        self.handler2.setLevel(logging.INFO)#Set verbosity of filehandler
        self.handler_format2 = logging.Formatter("%(message)-s")#Set Formatter
        self.handler2.setFormatter(self.handler_format2)#Add formatter to handler
        self.logger2.addHandler(self.handler2)#Add handler to logger2
        
        self.driver = webdriver.Chrome()
        self.driver.implicitly_wait(2)
        self.base_url = "http://"
        self.verificationErrors = []
        self.accept_next_alert = True
        
    
    def test_2(self):
        driver = self.driver
        
        
        def Close_popUps(current_handle):
            #Close any popup windows and return to main window
            for handle in self.driver.window_handles:
                if handle!=current_handle:
                    self.driver.switch_to_window(handle)
                    self.driver.close()
            self.driver.switch_to_window(current_handle)
        
        
        #row=0 #For Excel Manipulation of Cells -------------
        total_counter=0
        fail_counter=0
        L=list()# The list of urls that need revalidation after testing failure
        NoSitesLoaded=list()

        #Printing Blog to Excel-------------
        #row = row+1
        #_cell1 = ws.cell('%s%s'%(col1, row))
        #_cell2 = ws.cell('%s%s'%(col2, row))
        #_cell1.style.fill.fill_type = Fill.FILL_SOLID
        #_cell1.style.fill.start_color.index = 'FF9999'
        #_cell1.style.font.color.index = Color.WHITE
        #_cell2.style.font.color.index = Color.WHITE
        #_cell2.style.fill.fill_type = Fill.FILL_SOLID
        #_cell2.style.fill.start_color.index = 'FF9999'
        #_cell1.value = 'A/A'
        #_cell2.value = 'FIREFOX AUTOMATION TESTING'
        for url in lines:
            try:
                self.driver.get(url) #Go to the specified url
                main_window_ID=self.driver.current_window_handle#Save ID of main window
            except:
                fail_counter +=1#increase fail counter by1
                print('Error with url parser! Bypassing URL: %s')% ( self.driver.current_url)
                #Printing Blog to Excel-------------
                NoSitesLoaded.append(self.driver.current_url)#Add URL to revalidation list
                continue#return to for loop and run test for the next URL
            

            total_counter=total_counter+1
            self.logger2.propagate = False       
            self.logger2.info("===========================================================================")
            logging.info("======== %d Test STARTED for url:%s =========",total_counter,self.driver.current_url)
            #Printing Blog to Excel-------------
            #row = row+1
            #ws.cell('%s%s'%(col1, row)).value = '***********************************************************'
            #ws.cell('%s%s'%(col2, row)).value = '*********************#%s Test STARTED**************************************' % (total_counter)
            #End of Printing Blog to Excel-------------
            
            Close_popUps(main_window_ID)
            try:
                self.assertEqual("",driver.find_element_by_xpath("//link[contains(@href,'.css')]").text)#Check if js is inserted in html page source
                css_links=driver.find_elements(By.XPATH,"//link[contains(@href,'.css')]")
                #PRINT BLOCK
                #row= row+1
                #ws.cell('%s%s'%(col1, row)).value = '%d' % (total_counter)
                #ws.cell('%s%s'%(col2, row)).value = ' *****************#URL: %s*****************' % (self.driver.current_url)
                #END OF PRINT BLOCK
                             
                for links in css_links:
                    linkcss_text=links.get_attribute('href')
                    #print(linkcss_text)
                    
                    #PRINT BLOCK
                    #row= row+1
                    #_cell = ws.cell('%s%s'%(col2, row))
                    #_cell.style.fill.fill_type = Fill.FILL_SOLID
                    #_cell.style.fill.start_color.index = '99CCFF'
                    #ws.cell('%s%s'%(col1, row)).value = '%s%s' % (" ", " ")
                    #_cell.value = '#%s : CSS file'% (linkcss_text)
                    #print("Found jQuery")
                    #END OF PRINT BLOCK
                    logging.info("%s",linkcss_text)
                    continue
                    
                
            except:
                ###
                #row= row+1
                #ws.cell('%s%s'%(col1, row)).value = '%d' % (total_counter)
                #ws.cell('%s%s'%(col2, row)).value = ' *****************#URL: %s*****************' % (self.driver.current_url)
                #row= row+1
                #_cell = ws.cell('%s%s'%(col2, row))
                #_cell.style.fill.fill_type = Fill.FILL_SOLID
                #_cell.style.fill.start_color.index = 'B2FF66'
                #ws.cell('%s%s'%(col1, row)).value = '%s%s' % (" ", " ")
                #_cell.value = '#Not Found any css link'
                ###
                print("Not Found External css in url:"+self.driver.current_url)
                L.append(self.driver.current_url)#Add URL to revalidation list
                fail_counter +=1
                #logging.info("Not Found any css link")
                #logging.info("#%d Failed to display.",fail_counter)
                continue#return to for loop and run test for the next URL
            
        self.logger2.info("===========================================================================")
        self.logger2.info("===========================================================================")
        percentage_fail=(fail_counter/total_counter)*100#calculate percentage failure
        logging.warning("Total sites visited: %d",total_counter )#print total sites tested
        logging.warning("Sites CSS not appeared: %d, percentage: %d %%", fail_counter, percentage_fail)
        logging.warning("URLS needing revalidation: %s", L)#Print revalidation list
        logging.warning("URLS that do not work: %s", NoSitesLoaded)#Print revalidation list    
            
    def is_element_present(self, how, what):
        try: self.driver.find_element(by=how, value=what)
        except NoSuchElementException, e: return False
        return True
    
    def is_alert_present(self):
        try: self.driver.switch_to_alert()
        except NoAlertPresentException, e: return False
        return True
    
    def close_alert_and_get_its_text(self):
        try:
            alert = self.driver.switch_to_alert()
            alert_text = alert.text
            if self.accept_next_alert:
                alert.accept()
            else:
                alert.dismiss()
            return alert_text
        finally: self.accept_next_alert = True
    
    def tearDown(self):
        logging.shutdown()#Close all loggers and handlers
        self.driver.quit()#Close browser
        self.assertEqual([], self.verificationErrors)
        
if __name__ == "__main__":
    log_file = "reportExternalCSS.txt"
    f = open(log_file, "a")
    #Read testing urls from file urls.txt
    lines = open("urls.txt").read().splitlines()
    #Settings for Printing to Excel-------------
    """try:
        wb = load_workbook(filename = r'allcss.xlsx')
        ws = wb.create_sheet()
        ws.title = "Firefox JS"
    except:
        wb = Workbook()
        ws = wb.create_sheet()
        ws.title = "Firefox JS"
    dest_filename = r'allcss.xlsx'
    """
    try:
        unittest.main()
        #unittest.main()
    except Exception:
        pass
    finally:
        f.close()
        #wb.save(filename = dest_filename)#Save Finally to Excel-------------

    